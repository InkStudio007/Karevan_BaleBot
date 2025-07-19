from balethon import Client
from balethon.conditions import private, at_state, successful_payment
from balethon.objects import InlineKeyboard, InlineKeyboardButton, LabeledPrice
from balethon.errors.rpc_errors import ForbiddenError
from Validations import (
    validate_phone_number,
    validate_code_meli,
    validate_capacity,
    validate_price,
    validate_credit_card,
    validate_confirm
)
import os
from dotenv import load_dotenv
import json
import jdatetime
import shutil
import pandas
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from persiantools.digits import fa_to_en, ar_to_fa

#Variables

admin_ids = [403949029, 1828929996]
setting_payment_message_id = 0
signup_payment_message_ids = {}
CHANNEL_ID = 4858274378
excel_file_path = 'لیست مسافران کاروان.xlsx'
signup_json_file_path = os.path.abspath("JsonFiles/signup_datas.json")
payment_settings_json_file_path = os.path.abspath("JsonFiles/payment_settings_datas.json")
startpanel_informations_json_file_path = os.path.abspath("JsonFiles/startpanel_informations_datas.json")
userjoined_list_json_file_path = os.path.abspath("JsonFiles/userjoined_list.json")
User_SignUp_Data = {}


load_dotenv()

bot = Client(os.environ["TOKEN"])

# Json Files Structures

if os.path.exists(signup_json_file_path):
    with open(signup_json_file_path, "r", encoding="utf-8") as f:
        SignUp_Datas = json.load(f)
else:
    SignUp_Datas = {
        "Name": [],
        "Phone_Number": [],
        "Code_Meli": [],
        "BirthDate": [],
        "Photo_Filepath": []
    }

SignUp_Keys = ["Name", "Phone_Number", "Code_Meli", "BirthDate", "Photo_Filepath"]


if os.path.exists(payment_settings_json_file_path):
    with open(payment_settings_json_file_path, "r", encoding="utf-8") as f:
        Payment_Settings_Datas = json.load(f)
else:
    Payment_Settings_Datas = {
        "title": "",
        "description": "",
        "price": "",
        "credit_card": ""
    }

Payment_Settings_Keys = ["title", "description", "price", "credit_card"]


if os.path.exists(startpanel_informations_json_file_path):
    with open(startpanel_informations_json_file_path, "r", encoding="utf-8") as f:
        StartPanel_Informations_Datas = json.load(f)
else:
    StartPanel_Informations_Datas = {
        "description": "",
        "signup_capacity": 0,
        "signup_count": 0,
        "trip_is_start": False
    }

#Creating or Updating json files functions

def save_signup_data_to_json():
    with open(signup_json_file_path, "w", encoding="utf-8") as f:
        json.dump(SignUp_Datas, f, ensure_ascii=False, indent=2)

save_signup_data_to_json()


def save_payment_settings_data_to_json():
    with open(payment_settings_json_file_path, "w", encoding="utf-8") as f:
        json.dump(Payment_Settings_Datas, f, ensure_ascii=False, indent=2)

save_payment_settings_data_to_json()


def save_startpanel_informations_data_to_json():
    with open(startpanel_informations_json_file_path, "w", encoding="utf-8") as f:
        json.dump(StartPanel_Informations_Datas, f, ensure_ascii=False, indent=2)

save_startpanel_informations_data_to_json()


#Checking for Payment Settings


def payment_settings_check():    
    if ("" in (Payment_Settings_Datas["title"], Payment_Settings_Datas["description"], Payment_Settings_Datas["credit_card"], Payment_Settings_Datas["price"])):
        return False
    else:
        return True


#Checking admin and membership of chanel

def is_admin(user_id):
    global admin_ids
    return user_id in admin_ids

async def check_user_membership(user_id):
    try:
        member = await bot.get_chat_member(CHANNEL_ID, user_id)
        return member.status in ("member", "creator", "administrator")

    except Exception as e:
        print(f"Error checking user membership: {e}")
        return False


#Auto shutdown

async def auto_shutdown():
    global admin_ids
    for id in admin_ids:
        try:
            await bot.send_message(id, "ثبت نام پایان یافت سفر خوبی داشته باشید.")
        except Exception as e:
            print(f"❌ Failed to send message to {id}: {e}")
            
    StartPanel_Informations_Datas["trip_is_start"] = False
    save_startpanel_informations_data_to_json()


def persian_to_english_digits(text):
    fa_digit = ar_to_fa(text)
    en_digit = fa_to_en(fa_digit)
    return en_digit


#Commands

@bot.on_command(private)
async def admin_panel(*, message):
    global StartPanel_Informations_Datas

    if is_admin(user_id= message.author.id):
        if StartPanel_Informations_Datas["trip_is_start"]:
            await message.reply(
                "پنل مدیریت",
                InlineKeyboard(
                    [("اتمام ثبت نام.", "stop_signup")],
                    [("لیست مسافران.", "passengers_list")],
                    [("تعداد نفرات باقی مانده.", "remaining_capacity")],
                    [("حذف مسافر.", "remove_passenger")],
                    [("تنظیمات پرداخت.", "payment_settings")]
                )
            )
        else:
            await message.reply(
                "پنل مدیریت",
                InlineKeyboard(
                    [("شروع ثبت نام.", "start_signup")],
                    [("تنظیمات پرداخت.", "payment_settings")],
                    [("لیست مسافران.", "passengers_list")],
                    [("حذف مسافر.", "remove_passenger")]
                )
            )
    else:
        await message.reply("شما دسترسی به این دستور را ندارید.")

    User_SignUp_Data.pop(message.author.id, None)



@bot.on_command(private) 
async def start(*, message):
    await start_core(message, message.author.id)


async def start_core(message, user_id):
    if await check_user_membership(user_id):
        await message.reply(
            StartPanel_Informations_Datas["description"],
            InlineKeyboard(
                [("ثبت نام.", "SignUp")]
            )
        )    
    else:
        await message.reply(
            "سلام به بات ثبت نام در کاروان زیارتی خوش امدید"
            "برای ادامه ثبت نام اول عضو کانال شوید و بعد روی دکمه(عضو شدم)کلیک کنید.\n",
            InlineKeyboard(
                [InlineKeyboardButton('کانال کاروان', url='https://ble.ir/karevan_ziarati')],
                [('عضو شدم.', 'join')],
            )
        )

    message.author.set_state("")
    User_SignUp_Data.pop(user_id, None)


#CallBack Queryes

@bot.on_callback_query()
async def callback_handler(callback_query):
    global StartPanel_Informations_Datas, SignUp_Datas 
    user_id = callback_query.author.id

    callback_query.author.set_state("")

    #Admin Panel CallBacks

    if callback_query.data == "passengers_list":

        zip_path = "passports.zip"
        folder_path = "passport_photos"

        shutil.make_archive("passports", 'zip', folder_path)

        with open(signup_json_file_path, "r", encoding="utf-8") as f:
            json_SignUp_Datas = json.load(f)

        keys = list(json_SignUp_Datas.keys())
        keys_to_use = keys[:-1]

        filtered_dict = {k: json_SignUp_Datas[k] for k in keys_to_use}

        data_table = pandas.DataFrame(filtered_dict)        
        data_table.index += 1 
        data_table.columns = ['نام و نام خانوادگی', 'شماره تلفن', 'کد ملی', 'تاریخ تولد']

        data_table.to_excel(excel_file_path, index_label="ردیف")

        wb = load_workbook(excel_file_path)
        ws = wb.active

        max_col = ws.max_column
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 30

        max_row = ws.max_row
        for row_idx in range(1, max_row + 1):
            ws.row_dimensions[row_idx].height = 60

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = Font(size=26)

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        wb.save(excel_file_path)

        await bot.send_document(chat_id= callback_query.message.chat.id, document= open(excel_file_path, 'rb'))
        await bot.send_document(chat_id= callback_query.message.chat.id, document= open(zip_path, "rb"))
        await callback_query.answer("لیست مسافران در قالب فایل اکسل و عکس گذرنامه ها به صورت فشرده فرستاده شدند.")

        os.remove(zip_path)
        os.remove(excel_file_path)
        callback_query.author.set_state("")

    elif callback_query.data == "remove_passenger":        
        passenger_list = ""

        if StartPanel_Informations_Datas["signup_count"] > 0:
            for i, name in enumerate(SignUp_Datas["Name"]):
                passenger_list += f"{i + 1}. {name}\n"

            await callback_query.answer(f"لیست مسافران:\n\n{passenger_list}\n\nشماره مسافری که می‌خواهید حذف کنید را وارد کنید:")
            callback_query.author.set_state("REMOVE_PASSENGER_SELECT")

        else:
            await callback_query.answer("هنوز مسافری ثبت نام نکرده است")
            callback_query.author.set_state("")

    elif callback_query.data == "remaining_capacity":
        remaining_capacity = StartPanel_Informations_Datas["signup_capacity"] - StartPanel_Informations_Datas["signup_count"]

        await callback_query.answer(f"ظریفت باقی مانده: {remaining_capacity} نفر هست.")
        callback_query.author.set_state("")

    elif callback_query.data == "payment_settings":

        await bot.send_message(chat_id= callback_query.message.chat.id, text= "موضوع پرداخت را وارد کنید.")
        callback_query.author.set_state("TITLE")

    elif callback_query.data == "start_signup":
        if (payment_settings_check()):
            await bot.send_message(chat_id= callback_query.message.chat.id ,text= "توضیحات سفر را وارد کنید.")
            callback_query.author.set_state("TRIP_DESCRIPTION")

        else:
            await callback_query.answer("تنظیمات پرداخت روی هیچ مقداری تنظیم نشده است")
            callback_query.author.set_state("")

    elif callback_query.data == "stop_signup":
        StartPanel_Informations_Datas["trip_is_start"] = False
        save_startpanel_informations_data_to_json()

        await callback_query.answer("ثبت نام پایان یافت سفر خوبی داشته باشید.")
        callback_query.author.set_state("")


    #Start Panel CallBacks


    elif callback_query.data == "join":
        if await check_user_membership(user_id): 

            await bot.delete_message(callback_query.message.chat.id , callback_query.message.id)
            await callback_query.answer('شما عضو کانال هستید. حالا میتوانید برای ثبت نام اقدام کنید.')
            await start_core(callback_query.message, user_id)
            callback_query.author.set_state("")

        else:
            await callback_query.answer('شما عضو کانال نیستید. لطفاً ابتدا عضو کانال شوید.')
            callback_query.author.set_state("")

    elif callback_query.data == "SignUp":

        if (StartPanel_Informations_Datas["trip_is_start"]):
            User_SignUp_Data.pop(user_id, None)
            await bot.send_message(chat_id= callback_query.message.chat.id, text= "اسم و فامیلتون رو وارد کنید.")
            callback_query.author.set_state("NAME")

        else:
            await callback_query.answer("ثبت نام به پایان رسیده لطفا تا سفر بعد صبر کنید.")
            callback_query.author.set_state("")


    #paymant callbacks

    elif callback_query.data == "confirm_signup":

        user_data = User_SignUp_Data.get(user_id)

        if not user_data or len(user_data) < 6 or not user_data[5]:  
            await callback_query.answer("پس از پرداخت اقدام کنید")
            return

        # Delete old invoice 
        invoice_message_id = signup_payment_message_ids.pop(user_id, None)
        if invoice_message_id:
            try:
                await bot.delete_message(callback_query.message.chat.id, invoice_message_id)
            except ForbiddenError:
                print("⚠️ Bot was blocked or message not deletable.")
            except Exception as e:
                print(f"❌ Other error deleting message: {e}")

        # Save passport photo
        photo_path = f"passport_photos/{user_data[0]}_{user_data[2]}.jpg"
        with open(photo_path, "wb") as f:
            f.write(user_data[4])
        user_data[4] = photo_path

        for i in range(len(SignUp_Keys)):
            SignUp_Datas[SignUp_Keys[i]].append(user_data[i])

        StartPanel_Informations_Datas["signup_count"] += 1

        save_signup_data_to_json()
        save_startpanel_informations_data_to_json()

        User_SignUp_Data.pop(user_id, None)

        await callback_query.answer("ثبت نام با موفقیت کامل شد.")
        await bot.send_message(callback_query.message.chat.id, "اگر می‌خواهید دوستان یا آشنایان خود را ثبت نام کنید از دستور /start استفاده کنید.")

        if StartPanel_Informations_Datas["signup_count"] >= StartPanel_Informations_Datas["signup_capacity"]:
            await auto_shutdown()

        callback_query.author.set_state("")


    elif callback_query.data == "cancel_signup":

        invoice_message_id = signup_payment_message_ids.pop(user_id, None)
        if invoice_message_id:
            try:
                await bot.delete_message(callback_query.message.chat.id, invoice_message_id)
            except ForbiddenError:
                print("⚠️ Bot was blocked or message not deletable.")
            except Exception as e:
                print(f"❌ Other error deleting message: {e}")

        User_SignUp_Data.pop(user_id, None)

        await bot.send_message(callback_query.message.chat.id, "ثبت نام لغو شد. برای شروع مجدد /start را بزنید.")
        callback_query.author.set_state("")




# remove passengers state 

@bot.on_message(at_state("REMOVE_PASSENGER_SELECT"))
async def remove_passenger_state(message):
    try:
        index = int(persian_to_english_digits(message.text)) - 1
        if index < 0 or index >= len(SignUp_Datas["Name"]):
            raise IndexError

        for key in SignUp_Datas.keys():
            SignUp_Datas[key].pop(index)

        StartPanel_Informations_Datas["signup_count"] -= 1
        
        save_signup_data_to_json()
        save_startpanel_informations_data_to_json()

        await message.reply("مسافر با موفقیت حذف شد.")

    except (ValueError, IndexError):
        await message.reply("شماره وارد شده معتبر نیست. لطفاً دوباره تلاش کنید.")

    message.author.set_state("")

# Start Trip Information

@bot.on_message(at_state("TRIP_DESCRIPTION"))
async def trip_description_state(message):
    StartPanel_Informations_Datas["description"] = message.text

    await bot.send_message(chat_id= message.chat.id, text= "ظرفیت ثبت نام چند نفر هست؟")
    message.author.set_state("SIGNUP_CAPACITY")

@bot.on_message(at_state("SIGNUP_CAPACITY"))
async def SignUp_capacity_state(message):
    global SignUp_Datas

    if (validate_capacity(message.text)):

        StartPanel_Informations_Datas["signup_capacity"] = int(persian_to_english_digits(message.text))
        StartPanel_Informations_Datas["trip_is_start"] = True
        StartPanel_Informations_Datas["signup_count"] = 0
        SignUp_Datas = {
            "Name": [],
            "Phone_Number": [],
            "Code_Meli": [],
            "BirthDate": [],
            "Photo_Filepath": []
        }
        shutil.rmtree("passport_photos/")
        os.makedirs("passport_photos/")

        await bot.send_message(chat_id= message.chat.id, text= "ثبت نام با موفقیت اغاز شد.")

        save_startpanel_informations_data_to_json()
        save_signup_data_to_json()

        message.author.set_state("")

    else:
        await message.reply("مقدار واد شده یک عدد معتبر نمی باشد لطفا دوباره تلاش کنید.")


# Payment Settings

Payment_Settings_Data = []

@bot.on_message(at_state("TITLE"))
async def title_state(message):
    Payment_Settings_Data.append(message.text)

    await bot.send_message(chat_id= message.chat.id, text= "توضیحات پرداخت را وارد کنید.")
    message.author.set_state("DESCRIPTION")


@bot.on_message(at_state("DESCRIPTION"))
async def description_state(message):
    Payment_Settings_Data.append(message.text)

    await bot.send_message(chat_id= message.chat.id, text= "مبلغ را به ریال وارد کنید.")
    message.author.set_state("PRICE")


@bot.on_message(at_state("PRICE"))
async def price_state(message):
    if(validate_price(message.text)):
        Payment_Settings_Data.append(persian_to_english_digits(message.text))

        await bot.send_message(chat_id= message.chat.id, text= "شماره کارت را وارد کنید.")
        message.author.set_state("CREDIT_CARD")
    else:
        await message.reply("مبلغ وارد شده معتبر نیست لطفا دوباره تلاش کنید.")


@bot.on_message(at_state("CREDIT_CARD"))
async def credit_card_state(message):
    global setting_payment_message_id

    if (validate_credit_card(message.text)):
        Payment_Settings_Data.append(persian_to_english_digits(message.text))

        payment_message = await bot.send_invoice(
                chat_id=message.chat.id,
                title= Payment_Settings_Data[0],
                description= Payment_Settings_Data[1],
                payload=str(message.author.id),
                provider_token= Payment_Settings_Data[3],
                prices= [LabeledPrice(label="قیمت", amount= int(Payment_Settings_Data[2]))]
            )

        setting_payment_message_id = payment_message.id

        await bot.send_message(chat_id= message.chat.id, text= "تنظیمات پرداخت را تایید میکنید؟ (بله/خیر)")
        message.author.set_state("PAYMENT_CONFIRMATION")

    else:
        await message.reply("شماره کارت وارد شده معتبر نیسست لطفا دوباره تلاش کنید.")


@bot.on_message(at_state("PAYMENT_CONFIRMATION"))
async def payment_confirmation_state(message):
    global setting_payment_message_id

    if str(message.text).capitalize() in ("Yes", "No", "بله", "خیر"):

        if validate_confirm(message.text):

            await bot.delete_message(message.chat.id, setting_payment_message_id)

            for i in range(len(Payment_Settings_Keys)):
                Payment_Settings_Datas[Payment_Settings_Keys[i]] = Payment_Settings_Data[i]

            Payment_Settings_Data.clear()
            save_payment_settings_data_to_json()

            await message.reply("تنظیمات پرداخت با موفقیت ثبت شد.")

            message.author.set_state("") #reset state after confirmation
        else:

            await message.reply("دوباره با دستور /admin_panel تلاش کن.")
            message.author.set_state("") #reset state after no confirmation
    else:
        await message.reply("لطفا دوباره تلاش کن.")


# SignUp Process

@bot.on_message(at_state("NAME"))
async def name_state(message):
    User_SignUp_Data[message.author.id] = [message.text]
    await bot.send_message(chat_id= message.chat.id, text= "شماره تماس رو وارد کنید.")
    message.author.set_state("PHONE_NUMBER")


@bot.on_message(at_state("PHONE_NUMBER"))
async def phone_number_state(message):
    if validate_phone_number(message.text):
        User_SignUp_Data[message.author.id].append(persian_to_english_digits(message.text))

        await bot.send_message(chat_id= message.chat.id, text= "کد ملیتون رو وارد کنید.")
        message.author.set_state("CODE_MELI")
    else:
        await message.reply("شماره تلفن معتبر نیست دوباره تلاش کنید")


@bot.on_message(at_state("CODE_MELI"))
async def code_meli_state(message):
    if validate_code_meli(message.text):
        User_SignUp_Data[message.author.id].append(persian_to_english_digits(message.text))
        await bot.send_message(chat_id= message.chat.id, text= "تاریخ تولدت رو با فرمت 01-06-1367 وارد کن.")
        message.author.set_state("BIRTHDATE")
    else:
        await message.reply("کد ملیت معتبر نیست دوباره تلاش کن.")


@bot.on_message(at_state("BIRTHDATE"))
async def age_state(message):
    data_str = message.text.replace("/", "-")

    try:
        year, month, day = map(int, data_str.split("-"))
        shamsi_data = str(jdatetime.date(year, month, day))
        User_SignUp_Data[message.author.id].append(shamsi_data)

        await bot.send_message(message.chat.id, "عکسی واضح از صفحه اول گذرنامتون ارسال کنید")  
        message.author.set_state("PASSPORT")      

    except ValueError:
        await message.reply("تاریخ تولدتون معتبر نیست دوباره تلاش کنید.")


@bot.on_message(at_state("PASSPORT"))
async def passport_state(message):
    if message.photo:
        passport_photo = message.photo[-1]

        photo_file = await bot.download(passport_photo.id)
        User_SignUp_Data[message.author.id].append(photo_file)

        data = User_SignUp_Data[message.author.id]
        confirmation_message = (
            f"اسم و فامیلیتون: {data[0]}, "
            f"شماره تماستون: {data[1]}, "
            f"کد ملیتون: {data[2]}, "
            f"تاریخ تولدتون: {data[3]}\n"
            f"اطلاعاتتون درسته؟ (بله/خیر)"
        )
        await bot.send_message(chat_id=message.chat.id, text=confirmation_message)

        message.author.set_state("SIGNUP_CONFIRMATION")

    else:
        await bot.send_message(message.chat.id, "پیامی که ارسال کردی یه عکس نبود لطفا دوباره امتحان کن")


@bot.on_message(at_state("SIGNUP_CONFIRMATION"))
async def SignUp_Confirmation_state(message):
    if str(message.text).capitalize() in ("Yes", "No", "بله", "خیر"):
        if validate_confirm(message.text):
            await payment_state(message)
        else:
            await message.reply("میتوانید دوباره با دستور /start ثبت نام کنید.")
            message.author.set_state("")
            User_SignUp_Data.pop(message.author.id, None)
    else:
        await message.reply("متوجه نشدم, لطفا دوباره تلاش کنید.")


#Payment System

async def payment_state(message):
    await message.reply("در حال پردازش پرداخت...")

    payment_message = await bot.send_invoice(
        chat_id=message.chat.id,
        title= Payment_Settings_Datas['title'],
        description= Payment_Settings_Datas['description'],
        payload=str(message.author.id),
        provider_token= Payment_Settings_Datas['credit_card'],  
        prices=[LabeledPrice(label="قیمت", amount= int(Payment_Settings_Datas['price']))]
    )

    await bot.send_message(
        chat_id=message.chat.id,
        text="بعد از انجام پرداخت روی دکمه زیر کلیک کنید تا ثبت نام شما نهایی شود",
        reply_markup=InlineKeyboard(
            [("تکمیل ثبت‌نام", "confirm_signup")],
            [("لغو ثبت‌نام", "cancel_signup")]
        )
    )

    signup_payment_message_ids[message.author.id] = payment_message.id


@bot.on_message(successful_payment)
async def show_payment(message):
    try:
        user_id = int(message.successful_payment.invoice_payload)
        if user_id in User_SignUp_Data:
            User_SignUp_Data[user_id].append(True)  # mark payment as complete
    except Exception as e:
        print(f"⚠️ Payment error: {e}")
    

bot.run()
